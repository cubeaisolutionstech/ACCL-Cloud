"""
Proof of Calculation Utils

This module contains utility functions for creating detailed Excel proof of calculation
reports for budget vs sales analysis.
"""

import pandas as pd
import logging
from io import BytesIO
from typing import Optional, List

# Configure logging
logger = logging.getLogger(__name__)

# Branch mapping configuration
BRANCH_MAPPING = {
    'PONDY': 'PUDUCHERRY', 'PDY': 'PUDUCHERRY', 'Puducherry - PDY': 'PUDUCHERRY',
    'COVAI': 'COIMBATORE', 'CBE': 'COIMBATORE', 'Coimbatore - CBE': 'COIMBATORE',
    'ERD': 'ERODE', 'Erode - ERD': 'ERODE', 'ERD002': 'ERODE', 'ERD001': 'ERODE',
    'ERDTD1': 'ERODE', 'ERD003': 'ERODE', 'ERD004': 'ERODE', 'ERD005': 'ERODE',
    'ERD007': 'ERODE',
    'KRR': 'KARUR',
    'Chennai - CHN': 'CHENNAI', 'CHN': 'CHENNAI',
    'Tirupur - TPR': 'TIRUPUR', 'TPR': 'TIRUPUR',
    'Madurai - MDU': 'MADURAI', 'MDU': 'MADURAI',
    'POULTRY': 'POULTRY', 'Poultry - PLT': 'POULTRY',
    'SALEM': 'SALEM', 'Salem - SLM': 'SALEM',
    'HO': 'HO',
    'SLM002': 'SALEM', 'SLMTD1': 'SALEM',
    'BHV1': 'BHAVANI',
    'CBU': 'COIMBATORE',
    'VLR': 'VELLORE', 
    'TRZ': 'TRICHY', 
    'TVL': 'TIRUNELVELI',
    'NGS': 'NAGERCOIL', 
    'PONDICHERRY': 'PUDUCHERRY',
    'BLR': 'BANGALORE', 'BANGALORE': 'BANGALORE', 'BGLR': 'BANGALORE'
}


def create_proof_of_calculation_excel(budget_df: pd.DataFrame, 
                                     sales_df: pd.DataFrame, 
                                     selected_month: str,
                                     budget_exec_col: str,
                                     budget_area_col: str, 
                                     budget_sl_code_col: str,
                                     budget_product_group_col: str, 
                                     budget_qty_col: str,
                                     budget_value_col: str,
                                     sales_exec_col: str,
                                     sales_date_col: str,
                                     sales_area_col: str,
                                     sales_sl_code_col: str,
                                     sales_product_group_col: str,
                                     sales_qty_col: str,
                                     sales_value_col: str,
                                     selected_executives: Optional[List[str]],
                                     selected_branches: Optional[List[str]] = None) -> Optional[BytesIO]:
    """
    Creates detailed Excel file showing proof of calculation with proper branch mapping.
    
    Args:
        budget_df: Budget DataFrame
        sales_df: Sales DataFrame
        selected_month: Month in format 'MMM YY' (e.g., 'Jan 25')
        budget_exec_col: Budget executive column name
        budget_area_col: Budget area column name
        budget_sl_code_col: Budget SL code column name
        budget_product_group_col: Budget product group column name
        budget_qty_col: Budget quantity column name
        budget_value_col: Budget value column name
        sales_exec_col: Sales executive column name
        sales_date_col: Sales date column name
        sales_area_col: Sales area column name
        sales_sl_code_col: Sales SL code column name
        sales_product_group_col: Sales product group column name
        sales_qty_col: Sales quantity column name
        sales_value_col: Sales value column name
        selected_executives: List of selected executives
        selected_branches: Optional list of selected branches
    
    Returns:
        BytesIO buffer containing the Excel file, or None if error occurred
    """
    try:
        # Create copies to avoid modifying original DataFrames
        sales_df = sales_df.copy()
        budget_df = budget_df.copy()
        
        # STEP 1: Convert sales date and filter by month
        sales_df[sales_date_col] = pd.to_datetime(sales_df[sales_date_col], dayfirst=True, errors='coerce')
        filtered_sales_df = sales_df[sales_df[sales_date_col].dt.strftime('%b %y') == selected_month].copy()
        
        if filtered_sales_df.empty:
            logger.warning(f"No sales data found for {selected_month}")
            return None

        # STEP 2: Apply executive filtering
        if selected_executives:
            filtered_sales_df = filtered_sales_df[filtered_sales_df[sales_exec_col].isin(selected_executives)]
            budget_df = budget_df[budget_df[budget_exec_col].isin(selected_executives)]
        
        if filtered_sales_df.empty or budget_df.empty:
            logger.warning("No data found for selected executives")
            return None

        # STEP 3: Convert numeric columns
        filtered_sales_df[sales_value_col] = pd.to_numeric(filtered_sales_df[sales_value_col], errors='coerce').fillna(0)
        filtered_sales_df[sales_qty_col] = pd.to_numeric(filtered_sales_df[sales_qty_col], errors='coerce').fillna(0)
        budget_df[budget_value_col] = pd.to_numeric(budget_df[budget_value_col], errors='coerce').fillna(0)
        budget_df[budget_qty_col] = pd.to_numeric(budget_df[budget_qty_col], errors='coerce').fillna(0)
        
        # STEP 4: Store ORIGINAL values BEFORE any standardization for display
        budget_df['display_branch'] = budget_df[budget_area_col].copy()
        budget_df['display_exec'] = budget_df[budget_exec_col].copy()
        budget_df['display_sl_code'] = budget_df[budget_sl_code_col].copy()
        budget_df['display_product'] = budget_df[budget_product_group_col].copy()
        
        # STEP 5: Apply EXACT same standardization as main calculation
        # Standardize string columns
        filtered_sales_df[sales_area_col] = filtered_sales_df[sales_area_col].astype(str).str.strip()
        budget_df[budget_area_col] = budget_df[budget_area_col].astype(str).str.strip()
        filtered_sales_df[sales_product_group_col] = filtered_sales_df[sales_product_group_col].astype(str).str.strip()
        filtered_sales_df[sales_sl_code_col] = filtered_sales_df[sales_sl_code_col].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
        budget_df[budget_product_group_col] = budget_df[budget_product_group_col].astype(str).str.strip()
        budget_df[budget_sl_code_col] = budget_df[budget_sl_code_col].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)

        # CRITICAL FIX: Apply branch mapping EXACTLY as in main calculation
        # First extract the branch part after ' - ' and remove 'AAAA - ' prefix
        budget_df[budget_area_col] = budget_df[budget_area_col].str.split(' - ').str[-1].str.upper()
        budget_df[budget_area_col] = budget_df[budget_area_col].str.replace('AAAA - ', '', regex=False).str.upper()
        
        # Apply branch mapping
        budget_df[budget_area_col] = budget_df[budget_area_col].replace(BRANCH_MAPPING)
        filtered_sales_df[sales_area_col] = filtered_sales_df[sales_area_col].str.upper().replace(BRANCH_MAPPING)
        
        # Apply branch filtering
        if selected_branches:
            filtered_sales_df = filtered_sales_df[filtered_sales_df[sales_area_col].isin(selected_branches)].copy()
            budget_df = budget_df[budget_df[budget_area_col].isin(selected_branches)].copy()

        # Standardize product groups and SL codes
        filtered_sales_df[sales_product_group_col] = filtered_sales_df[sales_product_group_col].str.upper()
        filtered_sales_df[sales_sl_code_col] = filtered_sales_df[sales_sl_code_col].str.upper()
        budget_df[budget_product_group_col] = budget_df[budget_product_group_col].str.upper()
        budget_df[budget_sl_code_col] = budget_df[budget_sl_code_col].str.upper()

        # STEP 6: Create detailed proof of calculation
        logger.info("Creating detailed proof of calculation...")
        
        # Group budget data by Branch + SL Code + Product Group
        budget_grouped = budget_df.groupby([
            budget_area_col,
            budget_sl_code_col, 
            budget_product_group_col
        ]).agg({
            'display_exec': 'first',      
            'display_branch': 'first',    
            'display_sl_code': 'first',   
            'display_product': 'first',   
            budget_qty_col: 'sum',
            budget_value_col: 'sum'
        }).reset_index()
        
        logger.info(f"Budget records after grouping: {len(budget_grouped)}")
        
        # Initialize results storage
        final_results = []
        
        # Process each budget record
        for _, budget_row in budget_grouped.iterrows():
            standardized_branch = budget_row[budget_area_col]
            standardized_sl_code = budget_row[budget_sl_code_col]
            standardized_product = budget_row[budget_product_group_col]
            budget_qty = budget_row[budget_qty_col]
            budget_value = budget_row[budget_value_col]
            
            # Use display values for Excel output
            display_branch = budget_row['display_branch']
            display_exec = budget_row['display_exec']
            display_sl_code = budget_row['display_sl_code']
            display_product = budget_row['display_product']
            
            # Check if budget conditions are met
            budget_conditions_met = budget_qty > 0 and budget_value > 0
            
            if budget_conditions_met:
                # Find matching sales using standardized values
                matching_sales = filtered_sales_df[
                    (filtered_sales_df[sales_area_col] == standardized_branch) &
                    (filtered_sales_df[sales_sl_code_col] == standardized_sl_code) &
                    (filtered_sales_df[sales_product_group_col] == standardized_product)
                ]
                
                # Sum matching sales
                if not matching_sales.empty:
                    sales_qty_total = matching_sales[sales_qty_col].sum()
                    sales_value_total = matching_sales[sales_value_col].sum()
                else:
                    sales_qty_total = 0
                    sales_value_total = 0
                
                # Apply min logic (if sales > budget, use budget; else use sales)
                final_qty = min(budget_qty, sales_qty_total) if sales_qty_total > 0 else 0
                final_value = min(budget_value, sales_value_total) if sales_value_total > 0 else 0
                
                # Determine match status
                if sales_qty_total > 0 or sales_value_total > 0:
                    match_status = 'Mapped'
                else:
                    match_status = 'No Sales Data'
            else:
                final_qty = 0
                final_value = 0
                match_status = 'Budget Invalid'
            
            # Store result with original display values
            final_results.append({
                'Branch': display_branch,           
                'Executive Name': display_exec,     
                'SL Code': display_sl_code,         
                'Product Group': display_product,   
                'Budget Qty': round(budget_qty, 2),
                'Budget Value': round(budget_value, 2),
                'Sales Qty': round(final_qty, 2),
                'Sales Value': round(final_value, 2),
                'Standardized Branch': standardized_branch,  # For debugging
                'Match Status': match_status
            })
        
        # Convert to DataFrame and sort
        proof_df = pd.DataFrame(final_results)
        proof_df = proof_df.sort_values(['Standardized Branch', 'Executive Name', 'SL Code'])
        
        logger.info(f"Created proof with {len(proof_df)} records")
        
        # Create Excel file with formatting
        excel_buffer = BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            # Remove the standardized branch column before writing (it was just for debugging)
            output_df = proof_df.drop(columns=['Standardized Branch'])
            output_df.to_excel(writer, sheet_name='Proof of Calculation', index=False)
            
            # Apply formatting
            workbook = writer.book
            worksheet = writer.sheets['Proof of Calculation']
            
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Header formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Conditional formatting for Match Status
            from openpyxl.formatting.rule import CellIsRule
            
            green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
            
            # Apply to Match Status column (column I)
            match_status_col = 'I'
            last_row = len(output_df) + 1
            
            worksheet.conditional_formatting.add(f'{match_status_col}2:{match_status_col}{last_row}',
                CellIsRule(operator='equal', formula=['"Mapped"'], fill=green_fill))
            worksheet.conditional_formatting.add(f'{match_status_col}2:{match_status_col}{last_row}',
                CellIsRule(operator='equal', formula=['"No Sales Data"'], fill=red_fill))
            worksheet.conditional_formatting.add(f'{match_status_col}2:{match_status_col}{last_row}',
                CellIsRule(operator='equal', formula=['"Budget Invalid"'], fill=yellow_fill))
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add summary information
            worksheet.insert_rows(1, 3)
            
            total_records = len(output_df)
            mapped_records = len(output_df[output_df['Match Status'] == 'Mapped'])
            no_sales_records = len(output_df[output_df['Match Status'] == 'No Sales Data'])
            budget_invalid_records = len(output_df[output_df['Match Status'] == 'Budget Invalid'])
            
            worksheet['A1'] = f"Total Records: {total_records}"
            worksheet['A2'] = f"Mapped: {mapped_records} | No Sales Data: {no_sales_records} | Budget Invalid: {budget_invalid_records}"
            worksheet['A3'] = f"Generated for Month: {selected_month}"
            
            for i in range(1, 4):
                worksheet[f'A{i}'].font = Font(bold=True)
        
        excel_buffer.seek(0)
        return excel_buffer
        
    except Exception as e:
        logger.error(f"Error creating proof of calculation Excel: {e}")
        return None


def generate_proof_summary(proof_df: pd.DataFrame) -> dict:
    """
    Generate summary statistics from proof of calculation DataFrame.
    
    Args:
        proof_df: Proof of calculation DataFrame
    
    Returns:
        Dictionary containing summary statistics
    """
    if proof_df.empty:
        return {
            'total_records': 0,
            'mapped_records': 0,
            'no_sales_records': 0,
            'budget_invalid_records': 0,
            'mapping_percentage': 0.0
        }
    
    total_records = len(proof_df)
    mapped_records = len(proof_df[proof_df['Match Status'] == 'Mapped'])
    no_sales_records = len(proof_df[proof_df['Match Status'] == 'No Sales Data'])
    budget_invalid_records = len(proof_df[proof_df['Match Status'] == 'Budget Invalid'])
    mapping_percentage = (mapped_records / total_records * 100) if total_records > 0 else 0.0
    
    return {
        'total_records': total_records,
        'mapped_records': mapped_records,
        'no_sales_records': no_sales_records,
        'budget_invalid_records': budget_invalid_records,
        'mapping_percentage': round(mapping_percentage, 2)
    }


def validate_proof_data(budget_df: pd.DataFrame, 
                       sales_df: pd.DataFrame,
                       budget_exec_col: str,
                       budget_area_col: str,
                       budget_sl_code_col: str,
                       budget_product_group_col: str,
                       budget_qty_col: str,
                       budget_value_col: str,
                       sales_exec_col: str,
                       sales_date_col: str,
                       sales_area_col: str,
                       sales_sl_code_col: str,
                       sales_product_group_col: str,
                       sales_qty_col: str,
                       sales_value_col: str) -> tuple[bool, List[str]]:
    """
    Validate that all required columns exist in both DataFrames and perform data quality checks.
    
    Args:
        budget_df: Budget DataFrame
        sales_df: Sales DataFrame
        budget_exec_col: Budget executive column name
        budget_area_col: Budget area column name
        budget_sl_code_col: Budget SL code column name
        budget_product_group_col: Budget product group column name
        budget_qty_col: Budget quantity column name
        budget_value_col: Budget value column name
        sales_exec_col: Sales executive column name
        sales_date_col: Sales date column name
        sales_area_col: Sales area column name
        sales_sl_code_col: Sales SL code column name
        sales_product_group_col: Sales product group column name
        sales_qty_col: Sales quantity column name
        sales_value_col: Sales value column name
    
    Returns:
        Tuple of (is_valid, list_of_errors)
    """
    errors = []
    
    # Define required columns
    required_budget_cols = [
        budget_exec_col, budget_area_col, budget_sl_code_col, 
        budget_product_group_col, budget_qty_col, budget_value_col
    ]
    
    required_sales_cols = [
        sales_exec_col, sales_date_col, sales_area_col, sales_sl_code_col,
        sales_product_group_col, sales_qty_col, sales_value_col
    ]
    
    # Check if DataFrames are provided and not None
    if budget_df is None:
        errors.append("Budget DataFrame is None")
        return False, errors
    
    if sales_df is None:
        errors.append("Sales DataFrame is None")
        return False, errors
    
    # Check if DataFrames are empty
    if budget_df.empty:
        errors.append("Budget DataFrame is empty")
    
    if sales_df.empty:
        errors.append("Sales DataFrame is empty")
    
    # Check budget columns
    missing_budget_cols = [col for col in required_budget_cols if col not in budget_df.columns]
    if missing_budget_cols:
        errors.append(f"Missing budget columns: {missing_budget_cols}")
    
    # Check sales columns
    missing_sales_cols = [col for col in required_sales_cols if col not in sales_df.columns]
    if missing_sales_cols:
        errors.append(f"Missing sales columns: {missing_sales_cols}")
    
    # If basic validation passed, perform data quality checks
    if not missing_budget_cols and not missing_sales_cols and not budget_df.empty and not sales_df.empty:
        
        # Check for valid date format in sales data
        try:
            pd.to_datetime(sales_df[sales_date_col].head(10), dayfirst=True, errors='coerce')
        except Exception as e:
            errors.append(f"Invalid date format in sales date column '{sales_date_col}': {str(e)}")
        
        # Check if numeric columns contain valid numeric data
        try:
            pd.to_numeric(budget_df[budget_qty_col].head(10), errors='coerce')
        except Exception as e:
            errors.append(f"Invalid numeric data in budget quantity column '{budget_qty_col}': {str(e)}")
            
        try:
            pd.to_numeric(budget_df[budget_value_col].head(10), errors='coerce')
        except Exception as e:
            errors.append(f"Invalid numeric data in budget value column '{budget_value_col}': {str(e)}")
            
        try:
            pd.to_numeric(sales_df[sales_qty_col].head(10), errors='coerce')
        except Exception as e:
            errors.append(f"Invalid numeric data in sales quantity column '{sales_qty_col}': {str(e)}")
            
        try:
            pd.to_numeric(sales_df[sales_value_col].head(10), errors='coerce')
        except Exception as e:
            errors.append(f"Invalid numeric data in sales value column '{sales_value_col}': {str(e)}")
        
        # Check for completely null columns
        null_budget_cols = []
        for col in required_budget_cols:
            if budget_df[col].isnull().all():
                null_budget_cols.append(col)
        
        if null_budget_cols:
            errors.append(f"Budget columns contain only null values: {null_budget_cols}")
        
        null_sales_cols = []
        for col in required_sales_cols:
            if sales_df[col].isnull().all():
                null_sales_cols.append(col)
        
        if null_sales_cols:
            errors.append(f"Sales columns contain only null values: {null_sales_cols}")
        
        # Check if there are any valid budget records (qty > 0 and value > 0)
        try:
            budget_qty_numeric = pd.to_numeric(budget_df[budget_qty_col], errors='coerce').fillna(0)
            budget_value_numeric = pd.to_numeric(budget_df[budget_value_col], errors='coerce').fillna(0)
            valid_budget_records = ((budget_qty_numeric > 0) & (budget_value_numeric > 0)).sum()
            
            if valid_budget_records == 0:
                errors.append("No valid budget records found (records with qty > 0 and value > 0)")
        except Exception as e:
            errors.append(f"Error validating budget records: {str(e)}")
    
    is_valid = len(errors) == 0
    
    if not is_valid:
        logger.error(f"Validation failed with {len(errors)} errors: {errors}")
    else:
        logger.info("Data validation passed successfully")
    
    return is_valid, errors


def get_required_columns() -> tuple[List[str], List[str]]:
    """
    Get the standard list of required columns for budget and sales data.
    
    Returns:
        Tuple of (budget_columns, sales_columns)
    """
    budget_columns = [
        'Executive Name', 'Branch', 'SL Code', 'Product Group', 'Quantity', 'Value'
    ]
    
    sales_columns = [
        'Executive Name', 'Date', 'Branch', 'SL Code', 'Product Group', 'Quantity', 'Value'
    ]
    
    return budget_columns, sales_columns