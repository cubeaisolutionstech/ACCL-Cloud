# routes/combined_data.py - FIXED VERSION with Left Aligned First Column
from flask import Blueprint, request, jsonify
import pandas as pd
import numpy as np
from io import BytesIO
import traceback
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import logging
import base64
import re

# Create Blueprint
combined_bp = Blueprint('combined_data', __name__, url_prefix='/api/combined')

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@combined_bp.route('/combine-excel-files', methods=['POST'])
def combine_excel_files():
    """Combine multiple Excel files into one master file with specific sheet order and names - FIXED WITH LEFT ALIGNED FIRST COLUMN"""
    try:
        data = request.get_json()
        file_data_list = data.get('files', [])
        excel_formatting = data.get('excel_formatting', {})
        
        # FORCE title_option to 'none' to remove all titles
        title_option = 'none'  # FIXED: Always remove titles
        
        if not file_data_list:
            return jsonify({
                'success': False,
                'error': 'No files provided for combining'
            }), 400

        logger.info(f"Combining {len(file_data_list)} files with NO titles and FIRST COLUMN LEFT aligned")

        # Create a new workbook with specific sheet order
        combined_wb = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in combined_wb.sheetnames:
            combined_wb.remove(combined_wb['Sheet'])
        
        # Define the sheet order and names (matching frontend expectations)
        sheet_order = [
            "Sales Analysis Month wise",
            "Region wise analysis", 
            "Product wise analysis",
            "TS-PW",
            "ERO-PW"
        ]
        
        # Create empty sheets in the specified order
        for sheet_name in sheet_order:
            combined_wb.create_sheet(sheet_name)
        
        # FIXED: Define alignment styles
        data_alignment_left = Alignment(horizontal="left", vertical="center")      # For first column
        data_alignment_right = Alignment(horizontal="right", vertical="center")   # For other columns
        data_alignment_center = Alignment(horizontal="center", vertical="center") # For headers
        
        # Table header styles (keep center for headers only)
        table_header_font = Font(bold=True, size=11, color="FFFFFF")
        table_header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        table_header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Track sheets created and processing log
        sheets_created = []
        processing_log = []
        category_breakdown = {}

        # Process each file and distribute data to appropriate sheets
        for file_index, file_data in enumerate(file_data_list):
            try:
                file_content = file_data.get('content')
                file_name = file_data.get('name', f'file_{file_index}')
                category = file_data.get('metadata', {}).get('category', 'unknown')
                
                if not file_content:
                    logger.warning(f"No content for file {file_name}")
                    processing_log.append(f"⚠️ Skipped {file_name}: No content")
                    continue
                
                # Decode base64 content
                file_bytes = BytesIO(base64.b64decode(file_content))
                wb = openpyxl.load_workbook(file_bytes, data_only=False)
                
                logger.info(f"Processing {file_name} with {len(wb.sheetnames)} sheets")
                processing_log.append(f"✅ Processing {file_name}: {len(wb.sheetnames)} sheets")
                
                category_breakdown[category] = category_breakdown.get(category, 0) + len(wb.sheetnames)
                
                # Determine target sheet based on file content/category (matching frontend logic)
                target_sheet_name = None
                file_name_lower = file_name.lower()
                category_lower = category.lower()
                
                if ('sales' in file_name_lower or 'sales' in category_lower or 
                    'month' in file_name_lower):
                    target_sheet_name = sheet_order[0]  # Sales Analysis Monthwise
                elif ('region' in file_name_lower or 'region' in category_lower):
                    target_sheet_name = sheet_order[1]  # Region Wise Analysis
                elif ('product' in file_name_lower or 'product' in category_lower):
                    target_sheet_name = sheet_order[2]  # Product Wise Analysis
                elif ('ts' in file_name_lower or 'ts_pw' in category_lower or 
                      'ts-pw' in file_name_lower):
                    target_sheet_name = sheet_order[3]  # TS-PW
                elif ('ero' in file_name_lower or 'ero_pw' in category_lower or 
                      'ero-pw' in file_name_lower):
                    target_sheet_name = sheet_order[4]  # ERO-PW
                
                if not target_sheet_name:
                    logger.warning(f"No matching sheet for {file_name}")
                    processing_log.append(f"⚠️ No target sheet for {file_name}")
                    continue
                
                target_sheet = combined_wb[target_sheet_name]
                
                # Copy data from all sheets in the file to the target sheet
                for source_sheet_name in wb.sheetnames:
                    source_sheet = wb[source_sheet_name]
                    
                    # Find the next empty row in target sheet
                    max_row = target_sheet.max_row
                    if max_row == 1 and target_sheet.cell(row=1, column=1).value is None:
                        start_row = 1
                    else:
                        # FIXED: Smaller gap since no titles
                        start_row = max_row + 2  # Only 2 rows gap between sections
                    
                    # Copy data with formatting and apply proper alignment
                    for row_idx, row in enumerate(source_sheet.iter_rows(), 1):
                        is_header_row = False
                        
                        # Check if this is likely a header row
                        if row_idx <= 3:  # First 3 rows might be headers
                            row_values = [str(cell.value or '').upper() for cell in row if cell.value]
                            header_keywords = ['SALES', 'ORGANIZATION', 'BUDGET', 'ACTUAL', 'MONTH', 'VALUE', 'MT', 'NAME', 'TOTAL']
                            is_header_row = any(keyword in ' '.join(row_values) for keyword in header_keywords)
                        
                        for cell in row:
                            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                                continue
                                
                            if cell.value is not None:
                                new_cell = target_sheet.cell(
                                    row=cell.row + start_row - 1,
                                    column=cell.column,
                                    value=cell.value
                                )
                                
                                # Copy formatting
                                try:
                                    if cell.font:
                                        new_cell.font = Font(
                                            name=cell.font.name or 'Calibri',
                                            size=cell.font.size or 11,
                                            bold=cell.font.bold,
                                            italic=cell.font.italic,
                                            color=cell.font.color
                                        )
                                    
                                    if cell.fill and cell.fill.fill_type:
                                        new_cell.fill = PatternFill(
                                            fill_type=cell.fill.fill_type,
                                            start_color=cell.fill.start_color,
                                            end_color=cell.fill.end_color
                                        )
                                    
                                    # FIXED: Apply alignment based on column and row type
                                    if is_header_row:
                                        # Headers stay centered
                                        new_cell.alignment = table_header_alignment
                                        # Make headers bold and with header styling
                                        new_cell.font = Font(
                                            name=new_cell.font.name or 'Calibri',
                                            size=new_cell.font.size or 11,
                                            bold=True,
                                            color=new_cell.font.color or "000000"
                                        )
                                    else:
                                        # FIXED: Column A (first column) gets LEFT alignment, others get RIGHT
                                        if cell.column == 1:  # First column (A)
                                            new_cell.alignment = data_alignment_left
                                        else:  # All other columns
                                            new_cell.alignment = data_alignment_right
                                        
                                        # Apply number formatting for numeric values (except first column)
                                        if isinstance(cell.value, (int, float)) and cell.column > 1:
                                            if (excel_formatting and 
                                                excel_formatting.get('number_format', {}).get('apply_to_all_numbers', False)):
                                                format_pattern = excel_formatting.get('number_format', {}).get('format_pattern', '0.00')
                                                new_cell.number_format = format_pattern
                                            else:
                                                new_cell.number_format = '0.00'
                                    
                                    if cell.border:
                                        new_cell.border = Border(
                                            left=cell.border.left,
                                            right=cell.border.right,
                                            top=cell.border.top,
                                            bottom=cell.border.bottom
                                        )
                                    
                                    if cell.number_format:
                                        new_cell.number_format = cell.number_format
                                except Exception as e:
                                    logger.warning(f"Formatting copy error: {str(e)}")
                    
                    sheets_created.append({
                        'target_sheet': target_sheet_name,
                        'source_file': file_name,
                        'source_sheet': source_sheet_name,
                        'rows_added': source_sheet.max_row,
                        'title_option': 'none'
                    })
                    
                    logger.info(f"Added data from {file_name} to {target_sheet_name}")
                    
            except Exception as e:
                error_msg = f"Error processing file {file_name}: {str(e)}"
                logger.error(error_msg)
                processing_log.append(f"❌ {error_msg}")
                continue

        # Auto-adjust column widths for all sheets
        for sheet in combined_wb.worksheets:
            # Calculate column widths based on content
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                # Set width with appropriate limits
                if column_letter == 'A':
                    # Column A (product names) - wider for better readability
                    min_width = max(max_length + 2, 30)  # Increased minimum width
                    sheet.column_dimensions[column_letter].width = min(min_width, 60)
                else:
                    # Other columns
                    if max_length > 0:
                        adjusted_width = min(max_length + 3, 50)
                        sheet.column_dimensions[column_letter].width = max(adjusted_width, 12)
                    else:
                        sheet.column_dimensions[column_letter].width = 15

        # Freeze the first column in each sheet
        for sheet in combined_wb.worksheets:
            sheet.freeze_panes = 'B1'  # Freeze the first column (everything to the left of column B)

        # Set print settings for all sheets
        for sheet in combined_wb.worksheets:
            sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
            sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.fitToHeight = 0

        # Save combined workbook
        output = BytesIO()
        combined_wb.save(output)
        output.seek(0)
        combined_data = output.getvalue()

        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"Auditor_Format_{timestamp}.xlsx"

        # Generate metadata
        file_metadata = {
            'files_combined': len(file_data_list),
            'sheets_order': sheet_order,
            'total_sheets': len(sheet_order),
            'category_breakdown': category_breakdown,
            'generated_on': datetime.now().isoformat(),
            'file_size': len(combined_data),
            'source_files': [f.get('name') for f in file_data_list],
            'processing_log': processing_log,
            'sheets_details': sheets_created,
            'title_option_used': 'none',
            'alignment_used': 'first_column_left_others_right',
            'totalSourceFiles': len(file_data_list),
            'categoriesIncluded': list(category_breakdown.keys()),
            'generationType': 'auditor_format_clean'
        }

        logger.info(f"Successfully created combined file with {len(sheet_order)} sheets - FIRST COLUMN LEFT aligned, other columns RIGHT aligned, NO titles")

        return jsonify({
            'success': True,
            'file_data': base64.b64encode(combined_data).decode('utf-8'),
            'file_name': file_name,
            'metadata': file_metadata,
            'message': f'Successfully combined files into {len(sheet_order)} sheets with FIRST COLUMN LEFT aligned and NO titles'
        })

    except Exception as e:
        logger.error(f"Error combining Excel files: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': f'Failed to combine Excel files: {str(e)}'
        }), 500

# Export the blueprint
__all__ = ['combined_bp']
